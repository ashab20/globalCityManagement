import calendar
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from utils.database import Session
from models.shop_profile import ShopProfile
from models.bill_info import BillInfo
from models.bill_particular import BillParticular
from sqlalchemy.orm import joinedload
from sqlalchemy import func, extract
from decimal import Decimal
from fpdf import FPDF
import os
import subprocess
import calendar
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ShopOwnerDueReportView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.style = ttk.Style()
        self.configure_layout()
        self.create_report_view()
        self.load_report_data()

    def configure_layout(self):
        """Configure style and layout"""
        self.style.configure("TFrame", background="white")
        self.style.configure("TLabel", background="white")
        self.style.configure("Treeview", font=("Helvetica", 10), rowheight=25)
        self.style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))

    def create_report_view(self):
        """Create the report view with treeview and buttons"""
        # Create Treeview for the report
        columns = ["Shop", "Items"]
        # Add months columns dynamically
        self.month_columns = self.get_last_12_months()
        columns.extend(self.month_columns)
        columns.append("Total")
        
        self.tree = ttk.Treeview(
            self,
            bootstyle="primary",
            columns=columns,
            show="headings",
            height=15,
            style="ReportTree.Treeview"
        )
        
        # Configure columns
        col_widths = [120, 100] + [60] * len(self.month_columns) + [80]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor=tk.CENTER if col != "Items" else tk.W)

        # Add scrollbars
        yscroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        # Grid layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        # Configure grid weights
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Configure style
        style = ttk.Style()
        # style.configure("ReportTree.Treeview", rowheight=25, font=("Helvetica", 10))
        style.configure("ReportTree.Treeview",
            rowheight=25,
            font=("Helvetica", 10),
            bordercolor="gray",
            borderwidth=1,
            relief="solid"
        )

        
        # Add buttons frame
        button_frame = ttk.Frame(self)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")

        # Buttons
        ttk.Button(
            button_frame,
            text="Refresh",
            command=self.load_report_data,
            bootstyle="primary"
        ).pack(side="left", padx=5)

        ttk.Button(
            button_frame,
            text="Print Report",
            command=self.print_report,
            bootstyle="info"
        ).pack(side="left", padx=5)

        ttk.Button(
            button_frame,
            text="Export Excel",
            command=self.print_excel_report,
            bootstyle="success"
        ).pack(side="left", padx=5)

    def get_last_12_months(self):
        """Get the last 12 months in MMM-YY format"""
        today = datetime.now()
        months = []
        for i in range(6):
            # Calculate month and year
            month = today.month - i
            year = today.year
            if month <= 0:
                month += 6
                year -= 1
                
            # Format as MMM-YY
            month_name = calendar.month_abbr[month]
            formatted = f"{month_name}-{str(year)[2:]}"
            months.append(formatted)
            
        # Reverse to show most recent first
        return list(reversed(months))

    def load_report_data(self):
        """Load report data from database"""
        # Clear existing data
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        session = Session()
        try:
            # Get all shops with dues
            shops = session.query(ShopProfile).all()
            
            # Get distinct items (particulars)
            distinct_items = session.query(
                BillParticular.bill_particular.distinct().label("item")
            ).all()
            items = [item.item for item in distinct_items]
            
            # Create a data structure to store dues
            # Format: {shop_id: {item: {month_index: amount}}}
            report_data = {}
            
            # Initialize data structure
            for shop in shops:
                report_data[shop.id] = {
                    "name": f"{shop.shop_name} ({shop.shop_no})",
                    "items": {item: [0.0] * len(self.month_columns) for item in items}
                }
            
            # Get dues for each month
            for idx, month_col in enumerate(self.month_columns):
                # Extract month and year from column name (e.g., "Dec-24")
                month_name, year_short = month_col.split("-")
                year = 2000 + int(year_short)  # Convert to full year (e.g., 2024)
                
                # Map month name to number
                month_num = None
                for i in range(1, 13):
                    if calendar.month_abbr[i] == month_name:
                        month_num = i
                        break
                
                if not month_num:
                    continue
                
                # Query dues for this month
                dues = session.query(
                    ShopProfile.id,
                    BillParticular.bill_particular,
                    func.sum(BillParticular.due_amount).label("total_due")
                ).join(BillInfo, BillInfo.shop_id == ShopProfile.id)\
                 .join(BillParticular, BillParticular.bill_id == BillInfo.id)\
                 .filter(
                    extract('year', BillInfo.bill_date) == year,
                    extract('month', BillInfo.bill_date) == month_num,
                    BillParticular.due_amount > 0
                 ).group_by(ShopProfile.id, BillParticular.bill_particular).all()
                
                # Populate data structure
                for shop_id, item, due_amount in dues:
                    if shop_id in report_data and item in report_data[shop_id]["items"]:
                        report_data[shop_id]["items"][item][idx] = due_amount
            
            # Add data to treeview
            for shop in shops:
                if shop.id not in report_data:
                    continue
                    
                shop_data = report_data[shop.id]
                first_item = True
                
                for item, amounts in shop_data["items"].items():
                    # Calculate total for this item
                    total = sum(Decimal(a) for a in amounts)
                    
                    # Only show rows with non-zero total
                    if total == 0:
                        continue
                    
                    # Prepare row values
                    values = [shop_data["name"] if first_item else "", item]
                    values.extend([f"{amt:.2f}" if amt > 0 else "" for amt in amounts])
                    values.append(f"{total:.2f}")
                    
                    # Insert row
                    self.tree.insert("", "end", values=values)
                    first_item = False
                    
                # Add a separator row after each shop
                if not first_item:  # Only if we added items
                    self.tree.insert("", "end", values=[""] * len(values))
                    
        except Exception as e:
            Messagebox.show_error(f"Error loading report: {str(e)}", "Database Error")
            print(f"Error loading report: {str(e)}")
            traceback.print_exc()
        finally:
            session.close()

    def print_report(self):
        """Generate and print the detailed due report"""
        try:
            # Create PDF
            pdf = FPDF(orientation='L')  # Landscape orientation
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Add title
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Shop Owner Due Report", 0, 1, 'C')
            pdf.ln(5)
            
            # Get current date for report
            report_date = datetime.now().strftime("%B %d, %Y")
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 10, f"Report Date: {report_date}", 0, 1, 'C')
            pdf.ln(10)
            
            # Create table
            col_widths = [40, 40] + [25] * len(self.month_columns) + [30]
            total_width = sum(col_widths)
            
            # Table headers
            pdf.set_font("Arial", "B", 10)
            headers = ["Shop Name", "Items"] + self.month_columns + ["Total"]
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
            pdf.ln()
            
            # Table rows
            pdf.set_font("Arial", "", 9)
            for item in self.tree.get_children():
                values = self.tree.item(item)["values"]
                if not values or all(val == "" for val in values):
                    continue  # Skip empty separator rows
                
                for i, value in enumerate(values):
                    align = 'L' if i == 1 else 'C'  # Left-align items, center others
                    pdf.cell(col_widths[i], 8, str(value), 1, 0, align)
                pdf.ln()
            
            # Save and open
            filename = os.path.join(os.path.expanduser("~"), "Downloads", "shop_due_report.pdf")
            pdf.output(filename)
            
            if os.name == "nt":
                os.startfile(filename)
            elif os.name == "posix":
                subprocess.run(["open", filename])
            else:
                subprocess.run(["xdg-open", filename])
                
            Messagebox.show_info("Report saved and opened successfully!", "Success")
            
        except Exception as e:
            Messagebox.show_error(f"Error generating report: {str(e)}", "Print Error")
            traceback.print_exc()


    def print_excel_report(self):
        """Generate and save the due report as an Excel file"""
        try:
            from datetime import datetime
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Shop Due Report"

            # Headers
            headers = ["Shop Name", "Items"] + self.month_columns + ["Total"]
            ws.append(headers)

            # Style header
            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"].font = header_font
                ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")

            # Add rows from treeview
            for item in self.tree.get_children():
                values = self.tree.item(item)["values"]
                if not values or all(v == "" for v in values):
                    continue
                ws.append(values)

            # Auto column width
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            # Save file
            file_path = os.path.join(os.path.expanduser("~"), "Downloads", f"shop_due_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(file_path)

            # Open the file
            if os.name == "nt":
                os.startfile(file_path)
            elif os.name == "posix":
                subprocess.run(["open", file_path])
            else:
                subprocess.run(["xdg-open", file_path])

            Messagebox.show_info("Excel report saved and opened successfully!", "Success")

        except Exception as e:
            Messagebox.show_error(f"Error generating Excel report: {str(e)}", "Excel Error")
            traceback.print_exc()
