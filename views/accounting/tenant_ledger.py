from pydoc import text
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from utils.database import Session
from ttkbootstrap.dialogs import Messagebox
from models.acc_head_of_accounts import AccHeadOfAccounts
from datetime import datetime
from controllers.accounting_controller import AccountingController
from tkinter import StringVar
from models.shop_renter_profile import ShopRenterProfile
from sqlalchemy import text as sql_text
from tkinter import Toplevel, Text
import os
import subprocess
from fpdf import FPDF
from openpyxl import Workbook
from tkinter.filedialog import asksaveasfilename
from tkinter.messagebox import showinfo, showerror

class TenantLedgerView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.tenant_id = StringVar()
        self.from_date = StringVar()
        self.to_date = StringVar()
        self.head_mapping = {}
        self.tenant_mapping = {}

        # Configure styles
        style = ttk.Style()
        style.configure("TFrame", background="white")
        style.configure("TLabel", background="white")
        style.configure("TButton", font=("Helvetica", 10))
        style.configure("Treeview", rowheight=24, font=("Helvetica", 10))
        style.map("Treeview", background=[("selected", "#4361ee")], foreground=[("selected", "white")])

        # Main container
        self.main_container = ttk.Frame(self)
        self.main_container.pack(fill="both", expand=True, anchor="n")

        self.create_form()

    def create_form(self):
        """Creates the tenant ledger form."""
        form_frame = ttk.Frame(self.main_container)
        form_frame.pack(fill="x", pady=(4, 0), anchor="n")

        # Title
        title_label = ttk.Label(
            form_frame,
            text="Tenant Ledger",
            font=("Helvetica", 16, "bold"),
            bootstyle="primary"
        )
        title_label.pack(pady=(0, 6), anchor="center")

        # Input row
        input_row = ttk.Frame(form_frame)
        input_row.pack(fill="x", padx=4)

        label_style = {"bootstyle": "primary", "font": ("Helvetica", 10)}
        entry_style = {"font": ("Helvetica", 10)}

        # Tenant
        ttk.Label(input_row, text="Tenant:", **label_style).grid(row=0, column=0, padx=5, pady=(0, 2), sticky="w")
        self.tenant_combobox = ttk.Combobox(
            input_row,
            state="readonly",
            width=24,
            **entry_style,
            textvariable=self.tenant_id
        )
        self.tenant_combobox.grid(row=1, column=0, padx=5, pady=(0, 10), sticky="w")

        # From Date
        ttk.Label(input_row, text="From Date:", **label_style).grid(row=0, column=1, padx=5, pady=(0, 2), sticky="w")
        self.from_date_picker = ttk.DateEntry(
            input_row,
            dateformat="%Y-%m-%d",
            firstweekday=6,
            bootstyle="primary",
            width=20,
            startdate=None
        )
        self.from_date_picker.grid(row=1, column=1, padx=5, pady=(0, 10), sticky="w")

        # To Date
        ttk.Label(input_row, text="To Date:", **label_style).grid(row=0, column=2, padx=5, pady=(0, 2), sticky="w")
        self.to_date_picker = ttk.DateEntry(
            input_row,
            dateformat="%Y-%m-%d",
            firstweekday=6,
            bootstyle="primary",
            width=20,
            startdate=None
        )
        self.to_date_picker.grid(row=1, column=2, padx=5, pady=(0, 10), sticky="w")

        # Search Button
        self.search_button = ttk.Button(
            form_frame,
            text="Search",
            command=self.search_ledger,
            bootstyle="primary-outline",
            width=20
        )
        self.search_button.pack(pady=(0, 10), anchor="center")

        # Table frame inside bordered frame
        self.table_frame = ttk.Frame(self.main_container)
        self.table_frame.pack(fill="both", expand=True)

        self.load_tenant_list()

    def load_tenant_list(self):
        """Loads tenant list into the combobox."""
        try:
            session = Session()
            tenants = ShopRenterProfile.get_tenants(session)
            session.close()

            self.tenant_mapping = {
                f"{tenant.renter_name} (REF: {tenant.id})": tenant.id
                for tenant in tenants
            }
            self.tenant_combobox['values'] = list(self.tenant_mapping.keys())

        except Exception as e:
            Messagebox.show_error(message=f"Error loading tenants: {str(e)}", title="Error", parent=self)

    def search_ledger(self):
        """Handle ledger search and display results."""
        selected_tenant_display = self.tenant_combobox.get()
        from_date = self.from_date_picker.entry.get()
        to_date = self.to_date_picker.entry.get()

        selected_tenant_id = self.tenant_mapping.get(selected_tenant_display) if selected_tenant_display else None

        if not all([selected_tenant_id, from_date, to_date]):
            Messagebox.show_error(message="All fields are required!", title="Validation Error", parent=self)
            return

        try:
            session = Session()
            tenant_ledger_balance = session.execute(sql_text("""
                SELECT 
                    acc_head_of_accounts.head_name,
                    acc_head_of_accounts.id AS head_id,
                    teanant_trans_history.*,
                    shop_renter_profile.renter_name
                FROM teanant_trans_history
                INNER JOIN shop_renter_profile ON shop_renter_profile.id = teanant_trans_history.teanant_id
                INNER JOIN acc_head_of_accounts ON acc_head_of_accounts.id = teanant_trans_history.head_id
                WHERE shop_renter_profile.id = :tenant_id
                AND teanant_trans_history.trans_dt BETWEEN :from_date AND :to_date
                ORDER BY acc_head_of_accounts.id ASC
            """), {"tenant_id": selected_tenant_id, "from_date": from_date, "to_date": to_date}).fetchall()
            session.close()

            for widget in self.table_frame.winfo_children():
                widget.destroy()

            # Add border frame
            border_frame = ttk.Frame(self.table_frame, bootstyle="secondary", borderwidth=1, relief="solid")
            border_frame.pack(fill="both", expand=True, padx=10, pady=10)

            columns = ("Date", "Head Name", "Reference", "Debit", "Credit")
            tree = ttk.Treeview(
                border_frame, columns=columns, show="headings", height=12, bootstyle="primary"
            )
            tree.pack(fill="both", expand=True)

            tree.heading("Date", text="Date")
            tree.column("Date", anchor="center", width=100)

            tree.heading("Head Name", text="Head Name")
            tree.column("Head Name", anchor="w", width=150)

            tree.heading("Reference", text="Reference")
            tree.column("Reference", anchor="center", width=100)

            tree.heading("Debit", text="Debit")
            tree.column("Debit", anchor="e", width=80)

            tree.heading("Credit", text="Credit")
            tree.column("Credit", anchor="e", width=80)

            total_debit_balance = 0
            total_credit_balance = 0

            # Save data for print/export
            self.last_table_data = []
            for row in tenant_ledger_balance:
                date = row.trans_dt.strftime("%Y-%m-%d")
                head_name = row.head_name
                reference = row.id
                closing_balance = row.closing_amt
                amount = float(row.trans_amount)
                drcr_type = row.crdr_type

                debit_balance = float(closing_balance) + amount if drcr_type == 'dr' else 0
                credit_balance = float(closing_balance) - amount if drcr_type == 'cr' else 0

                self.last_table_data.append([
                    date, head_name, reference,
                    f"{debit_balance:,.2f}" if debit_balance else "0",
                    f"{credit_balance:,.2f}" if credit_balance else "0"
                ])

                tree.insert("", "end", values=(
                    date, head_name, reference,
                    f"{debit_balance:,.2f}" if debit_balance else "0",
                    f"{credit_balance:,.2f}" if credit_balance else "0"
                ))

                total_debit_balance += debit_balance
                total_credit_balance += credit_balance

            # Insert total row
            tree.insert("", "end", values=(
                "Total", "", "", 
                f"{total_debit_balance:,.2f}",
                f"{total_credit_balance:,.2f}"
            ))


            # Action Buttons Frame
            action_btn_frame = ttk.Frame(self.table_frame)
            action_btn_frame.pack(pady=5)

            ttk.Button(
                action_btn_frame,
                text="Print Preview",
                bootstyle="info-outline",
                command=lambda: self.print_preview(self.last_table_data, columns)
            ).pack(side="left", padx=10)


            ttk.Button(
                action_btn_frame,
                text="Export to Excel",
                bootstyle="success-outline",
                command=lambda: self.export_to_excel(self.last_table_data, columns)
            ).pack(side="left", padx=10)

            ttk.Button(
                action_btn_frame,
                text="Print PDF",
                bootstyle="danger-outline",
                command=self.print_ledger_pdf
            ).pack(side="left", padx=10)


        except Exception as e:
            Messagebox.show_error(message=f"Error searching ledger: {str(e)}", title="Error", parent=self)
    

    def export_to_excel(self, table_data, columns):
        try:
            # Ask user where to save
            file_path = asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save as Excel File"
            )

            if not file_path:
                return  # User cancelled

            wb = Workbook()
            ws = wb.active
            ws.title = "Tenant Ledger"

            # Write headers
            ws.append(columns)

            # Write data rows
            for row in table_data:
                ws.append(row)

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(file_path)
            showinfo("Export Successful", f"Excel file saved to:\n{file_path}")
        except Exception as e:
            showerror("Export Failed", f"Error: {str(e)}")


    def print_preview(self, table_data, columns):
        preview_win = ttk.Toplevel(self)
        preview_win.title("Print Preview")
        preview_win.geometry("700x500")

        # Optional: add scrollbars
        container = ttk.Frame(preview_win)
        container.pack(fill="both", expand=True)

        canvas = ttk.Canvas(container)
        scrollbar_y = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar_y.pack(side="right", fill="y")

        preview_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=preview_frame, anchor="nw")
        canvas.pack(side="left", fill="both", expand=True)

        preview_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.configure(yscrollcommand=scrollbar_y.set)

        # Headers
        for j, col in enumerate(columns):
            label = ttk.Label(preview_frame, text=col, font=("Helvetica", 10, "bold"), borderwidth=1, relief="solid", padding=5)
            label.grid(row=0, column=j, sticky="nsew")

        # Data rows
        for i, row in enumerate(table_data, start=1):
            for j, val in enumerate(row):
                label = ttk.Label(preview_frame, text=val, font=("Helvetica", 10), borderwidth=1, relief="solid", padding=5)
                label.grid(row=i, column=j, sticky="nsew")

        # Allow columns to expand
        for j in range(len(columns)):
            preview_frame.grid_columnconfigure(j, weight=1)

    def print_ledger_pdf(self):
        if not hasattr(self, 'last_table_data') or not self.last_table_data:
            Messagebox.show_error(title="No Data", message="Please perform a search before printing.", parent=self)
            return

        tenant_display = self.tenant_combobox.get().replace("/", "_").replace(" ", "_")
        filename = os.path.join(os.path.expanduser("~"), "Downloads", f"tenant_ledger_{tenant_display}.pdf")

        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            font_path = os.path.join("fonts", "NotoSansBengali-Regular.ttf")
            pdf.add_font("Noto", "", font_path, uni=True)
            pdf.set_font("Noto", "", 12)

            # Header
            pdf.cell(0, 10, "Tenant Ledger", ln=True, align="C")
            pdf.set_font("Noto", "", 10)
            pdf.cell(0, 10, f"Tenant: {self.tenant_combobox.get()}", ln=True)
            pdf.cell(0, 10, f"Period: {self.from_date_picker.entry.get()} to {self.to_date_picker.entry.get()}", ln=True)
            pdf.ln(5)

            # Table headers
            pdf.set_fill_color(220, 220, 220)
            pdf.set_font("Noto", "", 10)
            col_widths = [30, 50, 30, 30, 30]
            headers = ["Date", "Head", "Reference", "Debit", "Credit"]
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align="C", fill=True)
            pdf.ln()

            # Table rows
            for row in self.last_table_data:
                for i, cell in enumerate(row):
                    pdf.cell(col_widths[i], 8, str(cell), border=1, align="C")
                pdf.ln()

            # Save
            pdf.output(filename)

            # Open the PDF
            if os.name == 'posix':
                subprocess.run(['open', filename])
            elif os.name == 'nt':
                os.startfile(filename)
            else:
                subprocess.run(['xdg-open', filename])

            Messagebox.show_info(title="Print Success", message=f"PDF saved to Downloads:\n{filename}", parent=self)

        except Exception as e:
            Messagebox.show_error(title="Print Failed", message=str(e), parent=self)


