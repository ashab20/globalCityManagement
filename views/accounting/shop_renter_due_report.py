import calendar
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from utils.database import Session
from models.shop_profile import ShopProfile
from models.bill_info import BillInfo
from models.bill_particular import BillParticular
from sqlalchemy.orm import joinedload
from sqlalchemy import func, extract
from decimal import Decimal
from fpdf import FPDF
import os
import subprocess
import calendar
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ShopRenterDueReportView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.style = ttk.Style()
        self.configure_layout()
        self.create_report_view()
        self.load_report_data()

    def configure_layout(self):
        """Configure style and layout"""
        self.style.configure("TFrame", background="white")
        self.style.configure("TLabel", background="white")
        self.style.configure("Treeview", font=("Helvetica", 10), rowheight=25)
        self.style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"))

    def create_report_view(self):
        """Create the report view with treeview and buttons"""
        # Create Treeview for the report
        columns = ["Shop", "Items"]
        # Add months columns dynamically
        self.month_columns = self.get_last_12_months()
        columns.extend(self.month_columns)
        columns.append("Total")
        
        self.tree = ttk.Treeview(
            self,
            bootstyle="primary",
            columns=columns,
            show="headings",
            height=15,
            style="ReportTree.Treeview"
        )
        
        # Configure columns
        col_widths = [120, 100] + [60] * len(self.month_columns) + [80]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor=tk.CENTER if col != "Items" else tk.W)

        # Add scrollbars
        yscroll = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        # Grid layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        # Configure grid weights
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Configure style
        style = ttk.Style()
        # style.configure("ReportTree.Treeview", rowheight=25, font=("Helvetica", 10))
        style.configure("ReportTree.Treeview",
            rowheight=25,
            font=("Helvetica", 10),
            bordercolor="gray",
            borderwidth=1,
            relief="solid"
        )

        
        # Add buttons frame
        button_frame = ttk.Frame(self)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")

        # Buttons
        ttk.Button(
            button_frame,
            text="Refresh",
            command=self.load_report_data,
            bootstyle="primary"
        ).pack(side="left", padx=5)

        ttk.Button(
            button_frame,
            text="Print Report",
            command=self.print_report,
            bootstyle="info"
        ).pack(side="left", padx=5)

        ttk.Button(
            button_frame,
            text="Export Excel",
            command=self.print_excel_report,
            bootstyle="success"
        ).pack(side="left", padx=5)

    def get_last_12_months(self):
        """Get the last 12 months in MMM-YY format"""
        today = datetime.now()
        months = []
        for i in range(6):
            # Calculate month and year
            month = today.month - i
            year = today.year
            if month <= 0:
                month += 6
                year -= 1
                
            # Format as MMM-YY
            month_name = calendar.month_abbr[month]
            formatted = f"{month_name}-{str(year)[2:]}"
            months.append(formatted)
            
        # Reverse to show most recent first
        return list(reversed(months))

    def load_report_data(self):
        """Load due report grouped by Ownner → Shop with monthly totals and grand total"""
        from models.shop_allocation import ShopAllocation
        from models.shop_renter_profile import ShopRenterProfile
        from models.shop_owner_profile import ShopOwnerProfile

        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        session = Session()
        try:
            # Step 1: Load active shop-owner-shop mapping
            allocations = session.query(
                ShopOwnerProfile.id.label("owner_id"),
                ShopOwnerProfile.ownner_name,  # fixed typo here
                ShopProfile.id.label("shop_id"),
                ShopProfile.shop_name,
                ShopProfile.shop_no
            ).join(ShopProfile, ShopProfile.shop_owner_id == ShopOwnerProfile.id)\
            .join(ShopAllocation, ShopAllocation.shop_profile_id == ShopProfile.id)\
            .filter(ShopAllocation.close_status == 0)\
            .distinct().all()

            # Structure: {owner_id: {"ownner_name": str, "shops": [(shop_id, "Shop (No)")]}}
            owner_data = {}
            for row in allocations:
                owner = owner_data.setdefault(row.owner_id, {
                    "ownner_name": row.ownner_name,
                    "shops": []
                })
                owner["shops"].append((row.shop_id, f"{row.shop_name} ({row.shop_no})"))

            # Step 2: Load house rent dues per shop per month
            report_data = {}  # shop_id → [monthly values]
            for idx, month_col in enumerate(self.month_columns):
                month_name, year_short = month_col.split("-")
                year = 2000 + int(year_short)
                month_num = next((i for i in range(1, 13) if calendar.month_abbr[i] == month_name), None)
                if not month_num:
                    continue

                dues = session.query(
                    BillInfo.shop_id,
                    func.sum(BillParticular.due_amount).label("due")
                ).join(BillParticular, BillParticular.bill_id == BillInfo.id)\
                .join(ShopAllocation, ShopAllocation.shop_profile_id == BillInfo.shop_id)\
                .filter(
                    extract('year', BillInfo.bill_date) == year,
                    extract('month', BillInfo.bill_date) == month_num,
                    BillParticular.bill_type == "Bill",
                    BillParticular.bill_particular == "House Rent",
                    BillParticular.due_amount != None,
                    ShopAllocation.close_status == 0
                ).group_by(BillInfo.shop_id).all()

                for shop_id, due in dues:
                    report_data.setdefault(shop_id, [0.0] * len(self.month_columns))[idx] = float(due)

            # Step 3: Render rows
            grand_totals = [0.0] * len(self.month_columns)
            grand_total_amount = Decimal("0.0")

            for owner in owner_data.values():
                owner_inserted = False
                owner_totals = [0.0] * len(self.month_columns)

                for shop_id, shop_label in owner["shops"]:
                    dues = report_data.get(shop_id, [0.0] * len(self.month_columns))
                    row_total = sum(Decimal(str(v)) for v in dues)
                    if row_total == 0:
                        continue

                    # Add to owner total
                    for i, v in enumerate(dues):
                        owner_totals[i] += v

                    values = [owner["ownner_name"] if not owner_inserted else "", shop_label]
                    values.extend([f"{v:.2f}" if v else "" for v in dues])
                    values.append(f"{row_total:.2f}")
                    self.tree.insert("", "end", values=values)
                    owner_inserted = True

                # Owner total row
                if owner_inserted:
                    total_row = ["", "Total"]
                    total_row.extend([f"{v:.2f}" if v else "" for v in owner_totals])
                    total_sum = sum(Decimal(str(v)) for v in owner_totals)
                    total_row.append(f"{total_sum:.2f}")
                    self.tree.insert("", "end", values=total_row)

                    # Update grand total
                    for i, v in enumerate(owner_totals):
                        grand_totals[i] += v
                    grand_total_amount += total_sum

                    # Separator
                    self.tree.insert("", "end", values=[""] * (2 + len(self.month_columns)))

            # Grand total row at end
            total_row = ["", "Grand Total"]
            total_row.extend([f"{v:.2f}" if v else "" for v in grand_totals])
            total_row.append(f"{grand_total_amount:.2f}")
            self.tree.insert("", "end", values=total_row)

        except Exception as e:
            Messagebox.show_error(f"Error loading report: {str(e)}", "Database Error")
            traceback.print_exc()
        finally:
            session.close()



    def print_report(self):
        """Generate and print the detailed due report"""
        try:
            # Create PDF
            pdf = FPDF(orientation='L')  # Landscape orientation
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Add title
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "Due Report Shop Owner (Rent Only)", 0, 1, 'C')
            pdf.ln(5)
            
            # Get current date for report
            report_date = datetime.now().strftime("%B %d, %Y")
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 10, f"Report Date: {report_date}", 0, 1, 'C')
            pdf.ln(10)
            
            # Create table
            col_widths = [40, 40] + [25] * len(self.month_columns) + [30]
            total_width = sum(col_widths)
            
            # Table headers
            pdf.set_font("Arial", "B", 10)
            headers = ["Owner Name", "Shops"] + self.month_columns + ["Total"]
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
            pdf.ln()
            
            # Table rows
            pdf.set_font("Arial", "", 9)
            for item in self.tree.get_children():
                values = self.tree.item(item)["values"]
                if not values or all(val == "" for val in values):
                    continue  # Skip empty separator rows
                
                for i, value in enumerate(values):
                    align = 'L' if i == 1 else 'C'  # Left-align items, center others
                    pdf.cell(col_widths[i], 8, str(value), 1, 0, align)
                pdf.ln()
            
            # Save and open
            filename = os.path.join(os.path.expanduser("~"), "Downloads", "shop_due_report.pdf")
            pdf.output(filename)
            
            if os.name == "nt":
                os.startfile(filename)
            elif os.name == "posix":
                subprocess.run(["open", filename])
            else:
                subprocess.run(["xdg-open", filename])
                
            Messagebox.show_info("Report saved and opened successfully!", "Success")
            
        except Exception as e:
            Messagebox.show_error(f"Error generating report: {str(e)}", "Print Error")
            traceback.print_exc()


    def print_excel_report(self):
        """Generate and save the due report as an Excel file"""
        try:
            from datetime import datetime
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Shop Due Report"

            # Headers
            headers = ["Shop Name", "Items"] + self.month_columns + ["Total"]
            ws.append(headers)

            # Style header
            header_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}1"].font = header_font
                ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")

            # Add rows from treeview
            for item in self.tree.get_children():
                values = self.tree.item(item)["values"]
                if not values or all(v == "" for v in values):
                    continue
                ws.append(values)

            # Auto column width
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            # Save file
            file_path = os.path.join(os.path.expanduser("~"), "Downloads", f"shop_due_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(file_path)

            # Open the file
            if os.name == "nt":
                os.startfile(file_path)
            elif os.name == "posix":
                subprocess.run(["open", file_path])
            else:
                subprocess.run(["xdg-open", file_path])

            Messagebox.show_info("Excel report saved and opened successfully!", "Success")

        except Exception as e:
            Messagebox.show_error(f"Error generating Excel report: {str(e)}", "Excel Error")
            traceback.print_exc()
